"""
Gerador de Excel V4 - Sistema Insumos x Orçamento
Funções para geração otimizada de relatórios Excel
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from utils.fornecedor_utils import converter_nomenclatura

def criar_workbook_base() -> Workbook:
    """Cria workbook base com configurações padrão"""
    wb = Workbook()
    
    # Remover planilha padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    return wb


def configurar_estilos() -> Dict:
    """Define estilos padronizados para o Excel"""
    return {
        'header': {
            'font': Font(bold=True, color='FFFFFF', size=12),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'data': {
            'font': Font(size=10),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'currency': {
            'font': Font(size=10),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': 'R$ #,##0.00',
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'number': {
            'font': Font(size=10),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '#,##0.00',
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'title': {
            'font': Font(bold=True, size=14, color='1F4E79'),
            'alignment': Alignment(horizontal='center')
        },
        'subtitle': {
            'font': Font(bold=True, size=11, color='5B9BD5'),
            'alignment': Alignment(horizontal='left')
        }
    }


def obter_definicoes_colunas() -> List[Dict]:
    """
    VERSÃO MELHORADA - Define estrutura das colunas com novos campos

    Inclui: Mês, Ano de Apropriação e Classificação ABC (no final da tabela)
    """
    return [
        {'key': 'building_id', 'header': 'Building_ID', 'width': 12, 'style': 'data'},
        {'key': 'obra_nome', 'header': 'Obra', 'width': 40, 'style': 'data'},
        {'key': 'insumo_id', 'header': 'Insumo_ID', 'width': 15, 'style': 'data'},
        {'key': 'insumo_nome', 'header': 'Insumo', 'width': 50, 'style': 'data'},
        {'key': 'codigo_recurso', 'header': 'Código_Recurso', 'width': 18, 'style': 'data'},

        # Categoria e Grupo de Recursos
        {'key': 'categoria_recurso', 'header': 'Categoria_Recurso', 'width': 20, 'style': 'data'},
        {'key': 'grupo_recurso', 'header': 'Grupo_Recurso', 'width': 30, 'style': 'data'},

        {'key': 'categoria', 'header': 'Categoria', 'width': 20, 'style': 'data'},
        {'key': 'unidade', 'header': 'Unidade', 'width': 12, 'style': 'data'},

        # Tipo de documento com conversão de nomenclatura
        {'key': 'tipo_documento', 'header': 'Tipo_Documento', 'width': 18, 'style': 'data', 'convert': True},
        {'key': 'classificacao', 'header': 'Classificação', 'width': 18, 'style': 'data', 'convert': True},

        {'key': 'documento_origem', 'header': 'Documento_Origem', 'width': 20, 'style': 'data'},

        # Número de Medição
        {'key': 'numero_medicao', 'header': 'Número_Medição', 'width': 15, 'style': 'data'},

        # Fornecedor
        {'key': 'fornecedor', 'header': 'Fornecedor', 'width': 40, 'style': 'data'},

        {'key': 'building_unit_id', 'header': 'Building_Unit_ID', 'width': 18, 'style': 'data'},
        {'key': 'building_unit_name', 'header': 'Building_Unit_Name', 'width': 30, 'style': 'data'},
        {'key': 'apropriacao_completa', 'header': 'Apropriação_Completa', 'width': 40, 'style': 'data'},
        {'key': 'codigo_apropriacao', 'header': 'Código_Apropriação', 'width': 20, 'style': 'data'},

        # Colunas hierárquicas WBS
        {'key': 'nivel_1', 'header': 'Nível_1', 'width': 35, 'style': 'data'},
        {'key': 'nivel_2', 'header': 'Nível_2', 'width': 35, 'style': 'data'},
        {'key': 'nivel_3', 'header': 'Nível_3', 'width': 35, 'style': 'data'},
        {'key': 'nivel_4', 'header': 'Nível_4', 'width': 35, 'style': 'data'},

        {'key': 'quantidade', 'header': 'Quantidade', 'width': 15, 'style': 'number'},
        {'key': 'valor_unitario', 'header': 'Valor_Unitário', 'width': 18, 'style': 'currency'},
        {'key': 'valor_total', 'header': 'Valor_Total', 'width': 18, 'style': 'currency'},
        {'key': 'data_documento', 'header': 'Data', 'width': 12, 'style': 'data'},

        # Mês, Ano e Classificação ABC (no final para análise)
        {'key': 'mes_apropriacao', 'header': 'Mês_Apropriação', 'width': 16, 'style': 'data'},
        {'key': 'ano_apropriacao', 'header': 'Ano_Apropriação', 'width': 16, 'style': 'data'},
        {'key': 'classificacao_abc', 'header': 'Classificação_ABC', 'width': 18, 'style': 'data'},

        {'key': 'status', 'header': 'Status', 'width': 15, 'style': 'data'}
    ]


def criar_aba_lancamentos(wb: Workbook, lancamentos: List[Dict], estilos: Dict) -> None:
    """
    VERSÃO CORRIGIDA - Separa medição do documento_origem e processa fornecedores
    """
    import re
    
    ws = wb.create_sheet("Lançamentos")
    colunas = obter_definicoes_colunas()
    
    # Adicionar título
    total_colunas = len(colunas)
    ultima_coluna = get_column_letter(total_colunas)
    
    ws.merge_cells(f'A1:{ultima_coluna}1')
    ws['A1'] = f"Relatório de Lançamentos - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = estilos['title']['font']
    ws['A1'].alignment = estilos['title']['alignment']
    
    # Adicionar subtítulo
    ws.merge_cells(f'A2:{ultima_coluna}2')
    ws['A2'] = f"Total de {len(lancamentos)} lançamentos"
    ws['A2'].font = estilos['subtitle']['font']
    ws['A2'].alignment = estilos['subtitle']['alignment']
    
    # Linha em branco
    linha_atual = 4
    
    # Cabeçalhos
    for col_idx, coluna in enumerate(colunas, 1):
        cell = ws.cell(row=linha_atual, column=col_idx, value=coluna['header'])
        _aplicar_estilo(cell, estilos['header'])
        
        # Definir largura da coluna
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = coluna['width']
    
    linha_atual += 1
    
    # PROCESSAR DADOS COM SEPARAÇÃO DE MEDIÇÃO
    for lancamento in lancamentos:
        for col_idx, coluna in enumerate(colunas, 1):
            valor = lancamento.get(coluna['key'], '')
            
            # PROCESSAMENTO ESPECIAL PARA DOCUMENTO_ORIGEM
            if coluna['key'] == 'documento_origem' and valor:
                # Procurar por "Med." no texto
                match = re.search(r'Med\.?\s*0*(\d+)', str(valor), re.IGNORECASE)
                if match:
                    # Extrair número da medição
                    numero_medicao = int(match.group(1))
                    
                    # Atualizar o lançamento com o número da medição
                    lancamento['numero_medicao'] = numero_medicao
                    
                    # Remover a parte "Med.XXX" do documento_origem
                    valor = re.sub(r'\s*Med\.?\s*\d+', '', str(valor), flags=re.IGNORECASE).strip()
                    lancamento['documento_origem'] = valor
            
            # PROCESSAMENTO ESPECIAL PARA NUMERO_MEDICAO
            if coluna['key'] == 'numero_medicao':
                # Se for string vazia, deixar em branco
                if valor == '' or valor is None:
                    valor = ''
                else:
                    # Garantir que é inteiro
                    try:
                        valor = int(valor) if valor else ''
                    except (ValueError, TypeError):
                        valor = ''
            
            # Converter nomenclatura se necessário
            if coluna.get('convert') and valor:
                valor = converter_nomenclatura(str(valor))
            
            cell = ws.cell(row=linha_atual, column=col_idx, value=valor)
            _aplicar_estilo(cell, estilos[coluna['style']])
        
        linha_atual += 1
    
    # Criar tabela
    if lancamentos:
        total_colunas = len(colunas)
        ultima_coluna = get_column_letter(total_colunas)
        table_range = f"A4:{ultima_coluna}{linha_atual-1}"
        table = Table(displayName="TabelaLancamentos", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(table)


def criar_aba_resumo_obras(wb: Workbook, totais_por_obra: Dict, estilos: Dict) -> None:
    """Cria aba com resumo por obra"""
    ws = wb.create_sheet("Resumo por Obra")
    
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = "Resumo por Obra"
    ws['A1'].font = estilos['title']['font']
    ws['A1'].alignment = estilos['title']['alignment']
    
    linha_atual = 3
    
    # Cabeçalhos
    headers = ['Obra ID', 'Nome da Obra', 'Lançamentos', 'Valor Total', 'Valor Médio', 'Quantidade Total']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=linha_atual, column=col_idx, value=header)
        _aplicar_estilo(cell, estilos['header'])
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    linha_atual += 1
    
    # Dados ordenados por valor total (decrescente)
    obras_ordenadas = sorted(
        totais_por_obra.items(),
        key=lambda x: x[1]['total_valor'],
        reverse=True
    )
    
    total_geral_valor = 0
    total_geral_lancamentos = 0
    total_geral_quantidade = 0
    
    for obra_id, dados in obras_ordenadas:
        ws.cell(row=linha_atual, column=1, value=obra_id)
        ws.cell(row=linha_atual, column=2, value=dados['obra_nome'])
        ws.cell(row=linha_atual, column=3, value=dados['total_lancamentos'])
        
        cell_valor = ws.cell(row=linha_atual, column=4, value=dados['total_valor'])
        _aplicar_estilo(cell_valor, estilos['currency'])
        
        cell_medio = ws.cell(row=linha_atual, column=5, value=dados['valor_medio_lancamento'])
        _aplicar_estilo(cell_medio, estilos['currency'])
        
        cell_qtd = ws.cell(row=linha_atual, column=6, value=dados['total_quantidade'])
        _aplicar_estilo(cell_qtd, estilos['number'])
        
        total_geral_valor += dados['total_valor']
        total_geral_lancamentos += dados['total_lancamentos']
        total_geral_quantidade += dados['total_quantidade']
        
        linha_atual += 1
    
    # Linha de totais
    linha_atual += 1
    ws.cell(row=linha_atual, column=1, value="TOTAL GERAL").font = Font(bold=True)
    ws.cell(row=linha_atual, column=3, value=total_geral_lancamentos).font = Font(bold=True)
    
    cell_total_valor = ws.cell(row=linha_atual, column=4, value=total_geral_valor)
    cell_total_valor.font = Font(bold=True)
    _aplicar_estilo(cell_total_valor, estilos['currency'])
    
    cell_total_qtd = ws.cell(row=linha_atual, column=6, value=total_geral_quantidade)
    cell_total_qtd.font = Font(bold=True)
    _aplicar_estilo(cell_total_qtd, estilos['number'])


def criar_aba_categorias(wb: Workbook, totais_por_obra: Dict, estilos: Dict) -> None:
    """Cria aba com análise por categorias"""
    ws = wb.create_sheet("Análise por Categoria")
    
    # Consolidar categorias
    todas_categorias = {}
    for dados_obra in totais_por_obra.values():
        for categoria, info in dados_obra['categorias'].items():
            if categoria not in todas_categorias:
                todas_categorias[categoria] = {'count': 0, 'valor': 0}
            todas_categorias[categoria]['count'] += info['count']
            todas_categorias[categoria]['valor'] += info['valor']
    
    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = "Análise por Categoria"
    ws['A1'].font = estilos['title']['font']
    ws['A1'].alignment = estilos['title']['alignment']
    
    linha_atual = 3
    
    # Cabeçalhos
    headers = ['Categoria', 'Quantidade de Lançamentos', 'Valor Total', 'Percentual']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=linha_atual, column=col_idx, value=header)
        _aplicar_estilo(cell, estilos['header'])
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    
    linha_atual += 1
    
    # Calcular total para percentuais
    valor_total_geral = sum(info['valor'] for info in todas_categorias.values())
    
    # Dados ordenados por valor
    categorias_ordenadas = sorted(
        todas_categorias.items(),
        key=lambda x: x[1]['valor'],
        reverse=True
    )
    
    for categoria, info in categorias_ordenadas:
        percentual = (info['valor'] / valor_total_geral * 100) if valor_total_geral > 0 else 0
        
        ws.cell(row=linha_atual, column=1, value=categoria)
        ws.cell(row=linha_atual, column=2, value=info['count'])
        
        cell_valor = ws.cell(row=linha_atual, column=3, value=info['valor'])
        _aplicar_estilo(cell_valor, estilos['currency'])
        
        cell_perc = ws.cell(row=linha_atual, column=4, value=f"{percentual:.1f}%")
        _aplicar_estilo(cell_perc, estilos['data'])
        
        linha_atual += 1


def criar_aba_building_units(wb: Workbook, totais_por_obra: Dict, estilos: Dict) -> None:
    """Cria aba com análise por Building Units"""
    ws = wb.create_sheet("Building Units")
    
    # Consolidar Building Units
    todas_units = {}
    for dados_obra in totais_por_obra.values():
        for unit_id, info in dados_obra['building_units'].items():
            if unit_id not in todas_units:
                todas_units[unit_id] = {
                    'name': info['name'],
                    'count': 0,
                    'valor': 0
                }
            todas_units[unit_id]['count'] += info['count']
            todas_units[unit_id]['valor'] += info['valor']
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "Análise por Building Unit"
    ws['A1'].font = estilos['title']['font']
    ws['A1'].alignment = estilos['title']['alignment']
    
    linha_atual = 3
    
    # Cabeçalhos
    headers = ['Unit ID', 'Nome', 'Quantidade de Lançamentos', 'Valor Total', 'Percentual']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=linha_atual, column=col_idx, value=header)
        _aplicar_estilo(cell, estilos['header'])
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    
    linha_atual += 1
    
    # Calcular total para percentuais
    valor_total_geral = sum(info['valor'] for info in todas_units.values())
    
    # Dados ordenados por valor
    units_ordenadas = sorted(
        todas_units.items(),
        key=lambda x: x[1]['valor'],
        reverse=True
    )
    
    for unit_id, info in units_ordenadas:
        percentual = (info['valor'] / valor_total_geral * 100) if valor_total_geral > 0 else 0
        
        ws.cell(row=linha_atual, column=1, value=unit_id)
        ws.cell(row=linha_atual, column=2, value=info['name'])
        ws.cell(row=linha_atual, column=3, value=info['count'])
        
        cell_valor = ws.cell(row=linha_atual, column=4, value=info['valor'])
        _aplicar_estilo(cell_valor, estilos['currency'])
        
        cell_perc = ws.cell(row=linha_atual, column=5, value=f"{percentual:.1f}%")
        _aplicar_estilo(cell_perc, estilos['data'])
        
        linha_atual += 1

def criar_aba_metadados(wb: Workbook, configs: Dict, relatorio_processamento: Dict, estilos: Dict) -> None:
    """
    VERSÃO MELHORADA - Adiciona informações sobre novos campos
    
    SUBSTITUIR a função criar_aba_metadados original por esta
    """
    ws = wb.create_sheet("Metadados")
    
    # Título
    ws.merge_cells('A1:B1')
    ws['A1'] = "Metadados do Relatório"
    ws['A1'].font = estilos['title']['font']
    ws['A1'].alignment = estilos['title']['alignment']
    
    linha_atual = 3
    
    # Informações gerais
    metadados = [
        ('Data de Geração', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ('Sistema', 'Insumos x Orçamento V4 - Enhanced'),
        ('', ''),
        ('=== CONFIGURAÇÕES ===', ''),
        ('Usuário API', configs['api']['user']),
        ('Subdomínio', configs['api']['subdomain']),
        ('Cache Habilitado', 'Sim' if configs['cache']['enabled'] else 'Não'),
        ('Building Units Filter', 'Sim' if configs['building_units']['filter_enabled'] else 'Não'),
        ('Building Units Permitidos', ', '.join(map(str, configs['building_units']['allowed_ids']))),
        ('', ''),
        ('=== NOMENCLATURA ===', ''),
        ('APROPRIADO → INCORRIDO', 'Lançamentos executados/realizados'),
        ('PENDENTE → COMPROMETIDO', 'Lançamentos comprometidos/a realizar'),
        ('ORÇADO → ORÇADO', 'Mantém nomenclatura original'),
        ('', ''),
        ('=== NOVOS CAMPOS ===', ''),
        ('Número_Medição', 'Número da medição do contrato (1, 2, 3...)'),
        ('Fornecedor', 'Nome do fornecedor obtido via matching'),
        ('Categoria_Recurso', 'Categoria do recurso/insumo da API'),
        ('Grupo_Recurso', 'Grupo do recurso/insumo da API'),
        ('', ''),
        ('=== ESTATÍSTICAS ===', ''),
        ('Total de Obras', relatorio_processamento['resumo_geral']['total_obras']),
        ('Total de Lançamentos', relatorio_processamento['resumo_geral']['total_lancamentos']),
        ('Valor Total', f"R$ {relatorio_processamento['resumo_geral']['valor_total']:,.2f}"),
        ('Valor Médio por Obra', f"R$ {relatorio_processamento['resumo_geral']['valor_medio_por_obra']:,.2f}"),
        ('Lançamentos Médio por Obra', f"{relatorio_processamento['resumo_geral']['lancamentos_medio_por_obra']:.1f}"),
        ('', ''),
        ('=== ESTRUTURA HIERÁRQUICA ===', ''),
        ('Colunas Nível_1 a Nível_4', 'Hierarquia WBS com código e descrição'),
        ('Formato das Colunas', 'Código - Descrição (ex: 05.002 - Água Fria)'),
        ('Ordenação Recomendada', 'Use Nível_1, depois Nível_2, etc.'),
    ]
    
    # Adicionar top obras
    if 'top_obras_por_valor' in relatorio_processamento:
        metadados.append(('', ''))
        metadados.append(('=== TOP 5 OBRAS POR VALOR ===', ''))
        for i, (obra_id, valor, nome) in enumerate(relatorio_processamento['top_obras_por_valor'][:5], 1):
            metadados.append((f"{i}. Obra {obra_id}", f"R$ {valor:,.2f} - {nome}"))
    
    # Escrever metadados
    for chave, valor in metadados:
        if chave.startswith('==='):
            ws.cell(row=linha_atual, column=1, value=chave).font = Font(bold=True, color='5B9BD5')
        else:
            ws.cell(row=linha_atual, column=1, value=chave)
            ws.cell(row=linha_atual, column=2, value=valor)
        linha_atual += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50


def _aplicar_estilo(cell, style_dict: Dict) -> None:
    """Aplica estilo a uma célula"""
    for propriedade, valor in style_dict.items():
        setattr(cell, propriedade, valor)


def gerar_excel_completo(lancamentos: List[Dict], totais_por_obra: Dict, 
                        relatorio_processamento: Dict, configs: Dict, 
                        output_path: Path, logger) -> Path:
    """
    Gera Excel otimizado com hierarquia WBS (apenas abas essenciais)
    Returns: caminho do arquivo gerado
    """
    logger.info(f"Iniciando geração de Excel com {len(lancamentos)} lançamentos")
    
    # Criar workbook e estilos
    wb = criar_workbook_base()
    estilos = configurar_estilos()
    
    try:
        # Criar apenas abas essenciais
        criar_aba_lancamentos(wb, lancamentos, estilos)
        logger.info("Aba 'Lançamentos' criada")
        
        criar_aba_resumo_obras(wb, totais_por_obra, estilos)
        logger.info("Aba 'Resumo por Obra' criada")
        
        criar_aba_metadados(wb, configs, relatorio_processamento, estilos)
        logger.info("Aba 'Metadados' criada")
        
        # Salvar arquivo
        wb.save(output_path)
        logger.info(f"Excel salvo: {output_path}")
        
        # Verificar tamanho do arquivo
        tamanho_mb = output_path.stat().st_size / (1024 * 1024)
        logger.info(f"Tamanho do arquivo: {tamanho_mb:.2f} MB")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Erro ao gerar Excel: {str(e)}")
        raise


def gerar_nome_arquivo(timestamp: datetime = None) -> str:
    """Gera nome padrão para arquivo Excel"""
    if not timestamp:
        timestamp = datetime.now()
    
    return f"analise_insumos_orcamento_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"


def gerar_excel_simples(lancamentos: List[Dict], output_path: Path, logger) -> Path:
    """
    Gera Excel simples com apenas os lançamentos
    Returns: caminho do arquivo gerado
    """
    logger.info(f"Gerando Excel simples com {len(lancamentos)} lançamentos")
    
    wb = criar_workbook_base()
    ws = wb.create_sheet("Lançamentos")
    
    colunas = obter_definicoes_colunas()
    
    # Cabeçalhos
    for col_idx, coluna in enumerate(colunas, 1):
        ws.cell(row=1, column=col_idx, value=coluna['header'])
        ws.column_dimensions[get_column_letter(col_idx)].width = coluna['width']
    
    # Dados
    for row_idx, lancamento in enumerate(lancamentos, 2):
        for col_idx, coluna in enumerate(colunas, 1):
            valor = lancamento.get(coluna['key'], '')
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    wb.save(output_path)
    logger.info(f"Excel simples salvo: {output_path}")
    
    return output_path


if __name__ == '__main__':
    # Teste básico do gerador
    import sys
    from pathlib import Path
    
    # Adicionar src ao path
    sys.path.insert(0, str(Path(__file__).parent.parent))
    
    from config.settings import load_all_configs
    from utils.logger import get_main_logger
    
    configs = load_all_configs()
    logger = get_main_logger(configs)
    
    # Dados de teste
    lancamentos_teste = [
        {
            'building_id': 123,
            'obra_nome': 'Obra Teste',
            'insumo_nome': 'Cimento',
            'quantidade': 10.0,
            'valor_unitario': 50.0,
            'valor_total': 500.0,
            'categoria': 'Material'
        }
    ]
    
    totais_teste = {
        123: {
            'total_lancamentos': 1,
            'total_valor': 500.0,
            'obra_nome': 'Obra Teste',
            'categorias': {'Material': {'count': 1, 'valor': 500.0}},
            'building_units': {1: {'name': 'Direto', 'count': 1, 'valor': 500.0}}
        }
    }
    
    relatorio_teste = {
        'resumo_geral': {
            'total_obras': 1,
            'total_lancamentos': 1,
            'valor_total': 500.0,
            'valor_medio_por_obra': 500.0,
            'lancamentos_medio_por_obra': 1.0
        },
        'top_obras_por_valor': [(123, 500.0, 'Obra Teste')]
    }
    
    arquivo_teste = configs['paths']['reports'] / 'teste_excel.xlsx'
    gerar_excel_completo(lancamentos_teste, totais_teste, relatorio_teste, configs, arquivo_teste, logger)
    
    print("Gerador de Excel testado com sucesso!")